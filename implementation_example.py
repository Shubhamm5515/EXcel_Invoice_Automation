"""
Complete Implementation: OCR → JSON → Excel
Production-ready code for vehicle rental booking automation
"""

import json
import re
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class BookingDataExtractor:
    """Extract and normalize booking data from OCR and user text"""
    
    def __init__(self):
        self.confidence_weights = {
            'customer_name': 15,
            'mobile_number': 15,
            'vehicle_name': 10,
            'start_datetime': 15,
            'end_datetime': 15,
            'base_rent': 15,
            'total_amount': 15
        }
    
    def extract(self, ocr_text: str, user_text: str) -> Dict[str, Any]:
        """Main extraction method"""
        data = self._initialize_schema()
        
        # Extract from both sources
        self._extract_customer_info(ocr_text, user_text, data)
        self._extract_vehicle_info(user_text, data)
        self._extract_datetime_info(user_text, data)
        self._extract_pricing_info(user_text, data)
        self._extract_boolean_flags(user_text, data)
        
        # Validate and calculate
        self._verify_calculations(data)
        data['extraction_confidence'] = self._calculate_confidence(data)
        
        return data

    def _initialize_schema(self) -> Dict[str, Any]:
        """Initialize empty data schema"""
        return {
            "customer_name": None, "mobile_number": None, "company_name": None,
            "gstin": None, "address": None, "vehicle_name": None, "vehicle_number": None,
            "start_datetime": None, "end_datetime": None, "duration_days": None,
            "base_rent": None, "included_km": None, "extra_km": None,
            "extra_km_rate": None, "extra_km_charge": None, "extra_hour_rate": None,
            "extra_hours": None, "extra_hour_charge": None, "driver_allowance": None,
            "permit_charges": None, "parking_charges": None, "other_charges": None,
            "fuel_included": None, "toll_included": None, "pickup_drop_extra": None,
            "subtotal": None, "gst_rate": None, "gst_amount": None,
            "total_amount": None, "advance_paid": None, "balance_due": None,
            "calculation_verified": False, "calculation_notes": None,
            "extraction_confidence": "low", "notes": None
        }
    
    def _extract_customer_info(self, ocr_text: str, user_text: str, data: Dict):
        """Extract customer details"""
        combined_text = f"{ocr_text}\n{user_text}"
        
        # Customer name - multiple patterns
        name_patterns = [
            r'(?:Cx\s*name|customer\s*name|name)[:\s-]+([^\n]+)',
            r'Bill To:\s*([^\n]+)',
            r'Company[:\s]+([^\n]+)',
        ]
        for pattern in name_patterns:
            match = re.search(pattern, combined_text, re.IGNORECASE)
            if match:
                name = match.group(1).strip()
                # Clean up the name
                name = re.sub(r'[:\-]+$', '', name).strip()
                data['customer_name'] = name.title()
                data['company_name'] = name.title()
                break
        
        # GSTIN from OCR (more reliable)
        gstin_match = re.search(r'\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}\b', combined_text)
        if gstin_match:
            data['gstin'] = gstin_match.group(0)
        
        # Mobile - multiple patterns
        mobile_patterns = [
            r'(?:Cx\s*no|mobile|mob|ph|phone|contact)[:\s-]*(\d{10})',
            r'\b(\d{10})\b'
        ]
        for pattern in mobile_patterns:
            match = re.search(pattern, combined_text, re.IGNORECASE)
            if match:
                data['mobile_number'] = match.group(1)
                break
        
        # Address from OCR - improved patterns
        # Try multiple address patterns
        address_patterns = [
            # Pattern 1: Plot number with full address and pincode (most specific)
            r'(Plot\s+no\s+\d+[^,\n]+,[^,\n]+,\s*\d{6})',
            # Pattern 2: Plot number to pincode (greedy)
            r'(Plot\s+no\s+\d+.+?\d{6})',
            # Pattern 3: After "Address:" or "Office:"
            r'(?:Office|Address)[:\s]*([^\n]+\d{6})',
            # Pattern 4: Any line with state and pincode
            r'([^\n]*(?:Rajasthan|Delhi|Mumbai|Bangalore|Jaipur)[^\n]*\d{6})',
        ]
        
        for pattern in address_patterns:
            address_match = re.search(pattern, combined_text, re.IGNORECASE | re.DOTALL)
            if address_match:
                address = address_match.group(1).strip()
                # Clean up address - remove content after pincode
                address = re.sub(r'(\d{6}).*', r'\1', address, flags=re.DOTALL)
                # Remove extra spaces and newlines
                address = re.sub(r'\s+', ' ', address)
                # Remove leading/trailing punctuation
                address = address.strip('.,;: ')
                # Only set if it looks like a valid address (has pincode and reasonable length)
                if re.search(r'\d{6}', address) and len(address) > 20:
                    data['address'] = address
                    break

    def _extract_vehicle_info(self, user_text: str, data: Dict):
        """Extract vehicle details"""
        # Vehicle name - multiple patterns
        vehicle_patterns = [
            r'(?:Cat\s*type|vehicle|car)[:\s-]*([^\n\(]+)',
            r'(?:vehicle|car)[:\s-]*([^\n]+)',
        ]
        for pattern in vehicle_patterns:
            match = re.search(pattern, user_text, re.IGNORECASE)
            if match:
                vehicle = match.group(1).strip()
                # Remove vehicle number if included
                vehicle = re.sub(r'\([^)]+\)', '', vehicle).strip()
                data['vehicle_name'] = vehicle.title()
                break
        
        # Vehicle number - pattern like (RJ59CB2547)
        vehicle_num_match = re.search(r'\(([A-Z]{2}\d{2}[A-Z]{2}\d{4})\)', user_text, re.IGNORECASE)
        if vehicle_num_match:
            data['vehicle_number'] = vehicle_num_match.group(1).upper()
        
        # Running KM
        km_match = re.search(r'Running\s*km[:\s-]*(\d+)', user_text, re.IGNORECASE)
        if km_match:
            data['included_km'] = int(km_match.group(1))
    
    def _extract_datetime_info(self, user_text: str, data: Dict):
        """Extract and normalize dates/times"""
        # Pattern: 21jan to 23jan 2026 or 25/01/2026 7am to 31/01/2026 7am
        # Pattern 1: 21jan to 23jan 2026
        month_pattern = r'(\d{1,2})(\w{3})\s*to\s*(\d{1,2})(\w{3})\s*(\d{4})'
        match = re.search(month_pattern, user_text, re.IGNORECASE)
        
        if match:
            start_day, start_month, end_day, end_month, year = match.groups()
            
            # Convert month abbreviation to number
            months = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                     'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
            start_month_num = months.get(start_month.lower(), 1)
            end_month_num = months.get(end_month.lower(), 1)
            
            # Extract time
            time_pattern = r'Time[:\s-]*(\d{1,2}:\d{2}(?:am|pm)?)\s*to\s*(\d{1,2}:\d{2}(?:am|pm)?)'
            time_match = re.search(time_pattern, user_text, re.IGNORECASE)
            
            if time_match:
                start_time, end_time = time_match.groups()
                data['start_datetime'] = self._normalize_datetime_v2(
                    f"{start_day}/{start_month_num}/{year}", start_time
                )
                data['end_datetime'] = self._normalize_datetime_v2(
                    f"{end_day}/{end_month_num}/{year}", end_time
                )
            else:
                # Default times
                data['start_datetime'] = f"{year}-{start_month_num:02d}-{int(start_day):02d} 00:00"
                data['end_datetime'] = f"{year}-{end_month_num:02d}-{int(end_day):02d} 00:00"
        else:
            # Pattern 2: Standard date format
            datetime_pattern = r'(\d{1,2}[/.-]\d{1,2}[/.-]\d{4})\s*(\d{1,2}(?::\d{2})?(?:am|pm)?)\s*(?:to|till|-)\s*(\d{1,2}[/.-]\d{1,2}[/.-]\d{4})\s*(\d{1,2}(?::\d{2})?(?:am|pm)?)'
            match = re.search(datetime_pattern, user_text, re.IGNORECASE)
            
            if match:
                start_date, start_time, end_date, end_time = match.groups()
                data['start_datetime'] = self._normalize_datetime(start_date, start_time)
                data['end_datetime'] = self._normalize_datetime(end_date, end_time)
        
        # Calculate duration
        if data['start_datetime'] and data['end_datetime']:
            try:
                start = datetime.strptime(data['start_datetime'], '%Y-%m-%d %H:%M')
                end = datetime.strptime(data['end_datetime'], '%Y-%m-%d %H:%M')
                duration = (end - start).days
                if duration == 0:
                    duration = 1  # At least 1 day
                data['duration_days'] = duration
            except:
                pass
        
        # Fallback: duration mentioned separately
        duration_match = re.search(r'Duration[:\s-]*(\d+)\s*days?', user_text, re.IGNORECASE)
        if duration_match and not data['duration_days']:
            data['duration_days'] = int(duration_match.group(1))
    
    def _normalize_datetime_v2(self, date_str: str, time_str: str) -> str:
        """Convert to YYYY-MM-DD HH:MM format - version 2"""
        try:
            # Parse date (already in D/M/Y format)
            parts = date_str.split('/')
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            
            # Parse time
            time_str = time_str.lower().strip()
            hour, minute = 0, 0
            
            if 'am' in time_str or 'pm' in time_str:
                is_pm = 'pm' in time_str
                time_str = time_str.replace('am', '').replace('pm', '').strip()
                
                if ':' in time_str:
                    hour, minute = map(int, time_str.split(':'))
                else:
                    hour = int(time_str)
                
                if is_pm and hour != 12:
                    hour += 12
                elif not is_pm and hour == 12:
                    hour = 0
            else:
                if ':' in time_str:
                    hour, minute = map(int, time_str.split(':'))
                else:
                    hour = int(time_str) if time_str else 0
            
            return f"{year}-{month:02d}-{day:02d} {hour:02d}:{minute:02d}"
        except:
            return None
    
    def _normalize_datetime(self, date_str: str, time_str: str) -> str:
        """Convert to YYYY-MM-DD HH:MM format"""
        # Parse date
        date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y']
        parsed_date = None
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        
        if not parsed_date:
            return None
        
        # Parse time
        time_str = time_str.lower().strip()
        hour, minute = 0, 0
        
        if 'am' in time_str or 'pm' in time_str:
            is_pm = 'pm' in time_str
            time_str = time_str.replace('am', '').replace('pm', '').strip()
            
            if ':' in time_str:
                hour, minute = map(int, time_str.split(':'))
            else:
                hour = int(time_str)
            
            if is_pm and hour != 12:
                hour += 12
            elif not is_pm and hour == 12:
                hour = 0
        else:
            if ':' in time_str:
                hour, minute = map(int, time_str.split(':'))
            else:
                hour = int(time_str) if time_str else 0
        
        return parsed_date.replace(hour=hour, minute=minute).strftime('%Y-%m-%d %H:%M')

    def _extract_pricing_info(self, user_text: str, data: Dict):
        """Extract all pricing information"""
        # Base rent
        rent_patterns = [
            r'Rent[:\s-]*₹?\s*(\d+(?:,\d+)*)',
            r'(?:base\s*rent|rent)[:\s-]*₹?\s*(\d+(?:,\d+)*)',
        ]
        for pattern in rent_patterns:
            match = re.search(pattern, user_text, re.IGNORECASE)
            if match:
                data['base_rent'] = int(match.group(1).replace(',', ''))
                break
        
        # Security deposit
        security_match = re.search(r'Security[:\s-]*₹?\s*(\d+(?:,\d+)*)', user_text, re.IGNORECASE)
        if security_match:
            data['security_deposit'] = int(security_match.group(1).replace(',', ''))
        
        # Total amount
        total_patterns = [
            r'Total[:\s-]*₹?\s*(\d+(?:,\d+)*)',
            r'(?:total\s*amount)[:\s-]*₹?\s*(\d+(?:,\d+)*)',
        ]
        for pattern in total_patterns:
            match = re.search(pattern, user_text, re.IGNORECASE)
            if match:
                data['total_amount'] = int(match.group(1).replace(',', ''))
                break
        
        # Online received / Advance paid
        received_patterns = [
            r'Online\s*received[:\s-]*₹?\s*(\d+(?:,\d+)*)',
            r'(?:advance|paid)[:\s-]*₹?\s*(\d+(?:,\d+)*)',
        ]
        for pattern in received_patterns:
            match = re.search(pattern, user_text, re.IGNORECASE)
            if match:
                data['advance_paid'] = int(match.group(1).replace(',', ''))
                break
        
        # Pending amount
        pending_match = re.search(r'Pending\s*amount[:\s-]*₹?\s*(\d+(?:,\d+)*)', user_text, re.IGNORECASE)
        if pending_match:
            data['balance_due'] = int(pending_match.group(1).replace(',', ''))
        
        # Extra KM: 551×8:-4408 or 551x8=4408
        extra_km_pattern = r'(?:extra\s*km[^:]*)?(\d+)\s*[×x]\s*(\d+)[:\s=-]*(\d+)'
        extra_km_match = re.search(extra_km_pattern, user_text, re.IGNORECASE)
        if extra_km_match:
            data['extra_km'] = int(extra_km_match.group(1))
            data['extra_km_rate'] = int(extra_km_match.group(2))
            data['extra_km_charge'] = int(extra_km_match.group(3))
        
        # Extra KM rate (if not found above)
        if not data['extra_km_rate']:
            rate_match = re.search(r'extra\s*km[^0-9]*(\d+)[/\s]*km', user_text, re.IGNORECASE)
            if rate_match:
                data['extra_km_rate'] = int(rate_match.group(1))
        
        # Extra hour rate
        hour_rate_match = re.search(r'extra\s*hour[^0-9]*(\d+)[/\s]*hour', user_text, re.IGNORECASE)
        if hour_rate_match:
            data['extra_hour_rate'] = int(hour_rate_match.group(1))
    
    def _extract_boolean_flags(self, user_text: str, data: Dict):
        """Extract boolean inclusion flags"""
        text_lower = user_text.lower()
        
        # Fuel
        if 'fuel included' in text_lower or 'with fuel' in text_lower:
            data['fuel_included'] = True
        elif any(x in text_lower for x in ['fuel not included', 'without fuel', 'excluding fuel', 'fuel & toll:-exclude', 'fuel exclude']):
            data['fuel_included'] = False
        
        # Toll
        if 'toll included' in text_lower or 'with toll' in text_lower:
            data['toll_included'] = True
        elif any(x in text_lower for x in ['toll not included', 'without toll', 'excluding toll', 'fuel & toll:-exclude', 'toll exclude']):
            data['toll_included'] = False
        
        # Pickup/Drop
        if 'pickup drop extra' in text_lower or 'pickup/drop charges extra' in text_lower:
            data['pickup_drop_extra'] = True
        elif 'pickup drop included' in text_lower or 'with pickup drop' in text_lower:
            data['pickup_drop_extra'] = False

    def _verify_calculations(self, data: Dict):
        """Verify all calculations"""
        issues = []
        
        # Verify extra KM charge
        if data['extra_km'] and data['extra_km_rate']:
            calculated = data['extra_km'] * data['extra_km_rate']
            if data['extra_km_charge']:
                if abs(calculated - data['extra_km_charge']) > 5:
                    issues.append(f"Extra KM: Expected {calculated}, Found {data['extra_km_charge']}")
            else:
                data['extra_km_charge'] = calculated
        
        # Verify total
        if data['base_rent'] and data['total_amount']:
            calculated_total = data['base_rent']
            if data['extra_km_charge']:
                calculated_total += data['extra_km_charge']
            if data['extra_hour_charge']:
                calculated_total += data['extra_hour_charge']
            if data['driver_allowance']:
                calculated_total += data['driver_allowance']
            
            if abs(calculated_total - data['total_amount']) <= 10:
                data['calculation_verified'] = True
            else:
                issues.append(f"Total: Expected {calculated_total}, Found {data['total_amount']}")
                data['calculation_verified'] = False
        
        if issues:
            data['calculation_notes'] = '; '.join(issues)
    
    def _calculate_confidence(self, data: Dict) -> str:
        """Calculate extraction confidence score"""
        score = 0
        for field, weight in self.confidence_weights.items():
            if data.get(field):
                score += weight
        
        if score >= 80:
            return 'high'
        elif score >= 50:
            return 'medium'
        else:
            return 'low'


class ExcelWriter:
    """Write extracted data to Excel template"""
    
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.cell_map = {
            'B5': 'customer_name', 'B6': 'mobile_number', 'B7': 'company_name',
            'B8': 'gstin', 'B9': 'address', 'E5': 'vehicle_name',
            'B12': 'start_datetime', 'E12': 'end_datetime', 'G12': 'duration_days',
            'B15': 'base_rent', 'B16': 'included_km', 'B17': 'extra_km',
            'C17': 'extra_km_rate', 'D17': 'extra_km_charge',
            'C18': 'extra_hour_rate', 'B25': 'total_amount'
        }

    def write(self, data: Dict[str, Any], output_path: str) -> str:
        """Write data to Excel file"""
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        for cell_ref, field_name in self.cell_map.items():
            value = data.get(field_name)
            if value is None:
                continue
            
            cell = ws[cell_ref]
            
            # Apply formatting based on field type
            if field_name in ['base_rent', 'extra_km_charge', 'extra_hour_charge', 'total_amount']:
                cell.value = value
                cell.number_format = '₹#,##0.00'
            elif field_name in ['start_datetime', 'end_datetime']:
                cell.value = datetime.strptime(value, '%Y-%m-%d %H:%M')
                cell.number_format = 'DD/MM/YYYY HH:MM'
            elif isinstance(value, bool):
                cell.value = 'Yes' if value else 'No'
            else:
                cell.value = value
        
        wb.save(output_path)
        return output_path


# Example usage
if __name__ == '__main__':
    # Sample inputs
    ocr_text = """
    Bill To:
    Buen Manejo Del Campo India Pvt. Ltd.
    Office no.4, 2nd Floor, Anmol Pride,
    Baner, Pune - 411045
    GSTIN NO: 27AAHCB7551K1ZB
    """
    
    user_text = """
    Name- Buen manejo del Campo India pvt . Ltd
    Mobile - 8889302969
    Vehicle - Baleno
    Rent :-₹16200
    Kms-600km
    Extra km charged:-551×8:-4408
    Total:-20608
    Duration -6 days
    Start date and time - 25/01/2026 7am to 31/01/2026 7am
    Fuel and toll is not including
    Pickup drop charges extra
    Extra hour charge 300/hour
    Extra km charge 8/km
    """
    
    # Extract data
    extractor = BookingDataExtractor()
    extracted_data = extractor.extract(ocr_text, user_text)
    
    # Print JSON
    print(json.dumps(extracted_data, indent=2))
    
    # Write to Excel (uncomment when template exists)
    # writer = ExcelWriter('template.xlsx')
    # output_file = writer.write(extracted_data, 'output.xlsx')
    # print(f"Excel file created: {output_file}")
