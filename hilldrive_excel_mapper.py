"""
Hill Drive Invoice Template - Excel Automation
Customized for 'inn sample.xlsx' template structure
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from typing import Dict, Any, List, Optional
import re
import os
from PIL import Image
import io


class HillDriveExcelWriter:
    """Write booking data to Hill Drive invoice template"""
    
    def __init__(self, template_path: str = 'inn sample.xlsx', master_file: str = None):
        self.template_path = template_path
        self.master_file = master_file or 'generated_invoices/all_invoices.xlsx'
        
        # Cells that contain formulas - DO NOT OVERWRITE
        self.formula_cells = {
            'E18',  # Tax calculation
            'F23',  # Subtotal tax
            'G23',  # Subtotal amount
            'F27',  # Taxable amount (formula)
            'F28',  # CGST (formula)
            'F29',  # SGST (formula)
            'F35',  # Balance (formula)
        }
        
        # Cell mapping based on actual template structure from screenshot
        self.cell_map = {
            # Invoice header (Row 8)
            'invoice_number': 'C8',      # HD/2026-27/001 (ONLY THIS ONE)
            'invoice_number_old': 'D8',  # OLD number to clear
            'invoice_date': 'F8',        # 02/02/26
            
            # Customer details
            'customer_name': 'C10',      # NAME (Row 10) - Column C, not B
            'license_no': 'D10',         # License No. (Row 10)
            'address': 'C12',            # ADDRESS (Row 12) - Column C, not B11
            'phone_number': 'C14',       # PHONE NO. (Row 14) - Column C
            'place_of_supply': 'C15',    # PLACE OF SUPPLY (Row 15) - Column C
            
            # Delivery details
            'delivery_address': 'F11',   # DELIVERY ADD. (Row 11)
            'vehicle_number': 'F14',     # VEHICLE NO. (Row 14)
            
            # Service details (Row 18 - main service row)
            'service_name': 'A18',       # SERVICES (e.g., "Vehicle - Self Drive")
            'sac_code': 'B18',           # SAC code (996511)
            'no_of_days': 'C18',         # NO OF DAYS
            'quantity': 'D18',           # QTY (usually 1)
            # E18 has formula - don't write to it
            # F18, G18 have formulas - don't write to them
            
            # Additional charges
            'km_limit': 'B22',           # KILOMETER LIMIT (Row 22)
            
            # Right side amounts (Column F)
            'security_deposit': 'F25',   # Security Deposit (Row 25)
            'pickup_drop': 'F26',        # Pick & Drop Charges (Row 26)
            # F27, F28, F29 have formulas - don't write to them
            'igst': 'F30',               # IGST @ 5% (Row 30)
            
            # Payment details
            'payment_mode': 'B31',       # PAYMENT MODE (Row 31)
            'round_off': 'F31',          # Round Off (Row 31)
            'other_charges': 'F32',      # OTHER CHARGES (Row 32)
            'total_amount': 'F33',       # Total Amount (Row 33)
            'received_amount': 'F34',    # Received Amount (Row 34)
            # F35 has formula - don't write to it
            
            # Booking details
            'booking_datetime': 'B33',   # BOOKING DATE & TIME (Row 33)
        }
    
    def write(self, data: Dict[str, Any], output_path: str) -> str:
        """
        Write extracted booking data to Excel template
        
        Args:
            data: Extracted booking data dictionary
            output_path: Path to save the filled Excel file
            
        Returns:
            Path to the created file
        """
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Generate invoice number if not provided
        if not data.get('invoice_number'):
            data['invoice_number'] = self._generate_invoice_number()
        
        # Set invoice date
        if not data.get('invoice_date'):
            data['invoice_date'] = datetime.now().strftime('%d/%m/%y')
        
        # Fill the sheet with data
        self._fill_sheet_data(ws, data)
        
        # Embed document images if provided
        if data.get('document_images'):
            self._embed_document_images(ws, data['document_images'])
        
        # Save the file
        wb.save(output_path)
        return output_path
    
    def write_to_master(self, data: Dict[str, Any], sheet_name: str = None) -> Dict[str, str]:
        """
        Write booking data as a new sheet in the master Excel file
        
        Args:
            data: Extracted booking data dictionary
            sheet_name: Name for the new sheet (defaults to invoice number)
            
        Returns:
            Dictionary with master_file path and sheet_name
        """
        # Generate invoice number if not provided
        if not data.get('invoice_number'):
            data['invoice_number'] = self._generate_invoice_number()
        
        # Set invoice date
        if not data.get('invoice_date'):
            data['invoice_date'] = datetime.now().strftime('%d/%m/%y')
        
        # Use invoice number as sheet name if not provided
        if not sheet_name:
            sheet_name = data['invoice_number'].replace('/', '-')
        
        # Load or create master workbook
        if os.path.exists(self.master_file):
            # Load existing master file
            master_wb = openpyxl.load_workbook(self.master_file)
            print(f"📂 Loading existing master file: {self.master_file}")
        else:
            # Create new master file from template
            master_wb = openpyxl.load_workbook(self.template_path)
            # Rename the first sheet
            master_wb.active.title = sheet_name
            print(f"📂 Creating new master file: {self.master_file}")
        
        # Check if sheet name already exists
        if sheet_name in master_wb.sheetnames:
            # Add timestamp to make it unique
            timestamp = datetime.now().strftime('%H%M%S')
            sheet_name = f"{sheet_name}_{timestamp}"
        
        # Load template to copy from
        template_wb = openpyxl.load_workbook(self.template_path)
        template_ws = template_wb.active
        
        # Create new sheet in master workbook by copying template
        if len(master_wb.sheetnames) == 1 and master_wb.active.max_row == 1:
            # First sheet is empty, use it
            ws = master_wb.active
            ws.title = sheet_name
        else:
            # Copy template sheet to master workbook
            ws = master_wb.create_sheet(title=sheet_name)
            
            # Copy all cells from template
            for row in template_ws.iter_rows():
                for cell in row:
                    new_cell = ws[cell.coordinate]
                    
                    # Copy value
                    if cell.value:
                        new_cell.value = cell.value
                    
                    # Copy style
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # Copy row dimensions
            for row_num, row_dim in template_ws.row_dimensions.items():
                ws.row_dimensions[row_num].height = row_dim.height
            
            # Copy column dimensions
            for col_letter, col_dim in template_ws.column_dimensions.items():
                ws.column_dimensions[col_letter].width = col_dim.width
            
            # Copy merged cells
            for merged_range in template_ws.merged_cells.ranges:
                ws.merge_cells(str(merged_range))
        
        # Now fill the data (same as write method)
        self._fill_sheet_data(ws, data)
        
        # Embed document images if provided
        if data.get('document_images'):
            self._embed_document_images(ws, data['document_images'])
        
        # Save the master file
        master_wb.save(self.master_file)
        
        print(f"✅ Added sheet '{sheet_name}' to {self.master_file}")
        
        return {
            'master_file': self.master_file,
            'sheet_name': sheet_name,
            'invoice_number': data['invoice_number']
        }
    
    def _fill_sheet_data(self, ws, data: Dict[str, Any]):
        """Fill worksheet with booking data (extracted from write method)"""
        
        # DEBUG: Print extracted data
        print("\n" + "="*60)
        print("📊 DATA BEING WRITTEN TO EXCEL:")
        print("="*60)
        print(f"Customer Name: {data.get('customer_name')}")
        print(f"Company Name: {data.get('company_name')}")
        print(f"Address: {data.get('address')}")
        print(f"Mobile: {data.get('mobile_number')}")
        print(f"Phone: {data.get('phone_number')}")
        print("="*60 + "\n")
        
        # Fill customer details
        customer_name = data.get('customer_name') or data.get('company_name')
        if customer_name:
            self._set_cell(ws, 'customer_name', customer_name)
            print(f"✅ Set customer_name to C10: {customer_name}")
        else:
            print("⚠️  WARNING: No customer name found!")
        
        # Clear the old invoice number in D8
        try:
            ws['D8'].value = None  # Clear old invoice number
            print("✅ Cleared old invoice number in D8")
        except:
            pass
        
        # Format address properly - handle multiline
        address = data.get('address')
        if address:
            # If address is too long, it might need wrapping
            self._set_cell(ws, 'address', address)
            # Enable text wrapping for address cell
            from openpyxl.styles import Alignment
            ws[self.cell_map['address']].alignment = Alignment(wrap_text=True, vertical='top')
            print(f"✅ Set address to C12: {address[:50]}...")
        else:
            print("⚠️  WARNING: No address found!")
        
        # Try both mobile_number and phone_number
        phone = data.get('mobile_number') or data.get('phone_number')
        if phone:
            self._set_cell(ws, 'phone_number', phone)
            print(f"✅ Set phone to C14: {phone}")
        else:
            print("⚠️  WARNING: No phone number found!")
        
        self._set_cell(ws, 'place_of_supply', data.get('place_of_supply', 'Jaipur'))
        
        # Fill delivery details
        self._set_cell(ws, 'delivery_address', data.get('delivery_address', 'Jaipur hub'))
        self._set_cell(ws, 'vehicle_number', data.get('vehicle_number'))
        
        # Fill service details
        service_name = self._format_service_name(data)
        self._set_cell(ws, 'service_name', service_name)
        self._set_cell(ws, 'sac_code', '996511')  # Standard SAC for car rental
        self._set_cell(ws, 'no_of_days', data.get('duration_days'))
        self._set_cell(ws, 'quantity', 1)
        
        # Calculate amounts
        if data.get('total_amount'):
            final_total = data.get('total_amount')
            security = data.get('security_deposit') or 0
            taxable_amount = final_total - security
            amount_with_gst = taxable_amount
        else:
            taxable_amount = self._calculate_taxable_amount(data)
            amount_with_gst = taxable_amount * 1.05
        
        # Write to G18 - this will trigger all formula calculations
        ws['G18'] = amount_with_gst
        
        # Additional details
        self._set_cell(ws, 'km_limit', data.get('included_km') or 0)
        self._set_cell(ws, 'security_deposit', data.get('security_deposit') or 0)
        self._set_cell(ws, 'pickup_drop', data.get('pickup_drop_charges') or 0)

        # Payment details
        self._set_cell(ws, 'payment_mode', data.get('payment_mode') or 'Online')
        self._set_cell(ws, 'other_charges', data.get('other_charges') or 0)
        self._set_cell(ws, 'round_off', 0)
        self._set_cell(ws, 'igst', 0)  # IGST is 0 for intra-state
        
        # Total amount and received amount
        if data.get('total_amount'):
            self._set_cell(ws, 'total_amount', data.get('total_amount'))
        else:
            total = amount_with_gst + (data.get('security_deposit') or 0)
            self._set_cell(ws, 'total_amount', round(total, 2))
        
        self._set_cell(ws, 'received_amount', data.get('advance_paid') or 0)
        
        # Booking date & time
        booking_dt = self._format_booking_datetime(data)
        self._set_cell(ws, 'booking_datetime', booking_dt)
        
        # Invoice details
        self._set_cell(ws, 'invoice_number', data['invoice_number'])
        self._set_cell(ws, 'invoice_date', data['invoice_date'])
    
    def _set_cell(self, ws, field_name: str, value):
        """Set cell value if field exists in mapping and is not a formula cell"""
        if field_name in self.cell_map and value is not None:
            cell_ref = self.cell_map[field_name]
            # Don't overwrite formula cells
            if cell_ref not in self.formula_cells:
                try:
                    # Check if cell is part of a merged range
                    cell = ws[cell_ref]
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        # Find the top-left cell of the merged range
                        for merged_range in ws.merged_cells.ranges:
                            if cell_ref in merged_range:
                                # Get the top-left cell of the merged range
                                top_left = merged_range.start_cell
                                ws[top_left.coordinate] = value
                                return
                    else:
                        ws[cell_ref] = value
                except Exception as e:
                    print(f"⚠️  Failed to set {field_name} at {cell_ref}: {e}")
            else:
                print(f"⚠️  Skipping {cell_ref} - contains formula")
    
    def _format_service_name(self, data: Dict) -> str:
        """Format service name from vehicle and rental type"""
        vehicle = data.get('vehicle_name', 'Vehicle')
        rental_type = data.get('rental_type', 'Self Drive')
        return f"{vehicle} - {rental_type}"
    
    def _calculate_taxable_amount(self, data: Dict) -> float:
        """
        Calculate taxable amount (before GST)
        This is the base amount on which GST will be calculated
        """
        # If we have base_rent and extra charges, sum them up
        base = data.get('base_rent') or 0
        extra_km = data.get('extra_km_charge') or 0
        extra_hour = data.get('extra_hour_charge') or 0
        other = data.get('other_charges') or 0
        
        # Calculate subtotal
        subtotal = base + extra_km + extra_hour + other
        
        # If subtotal is 0, try to extract from total_amount
        if subtotal == 0 and data.get('total_amount'):
            total = data.get('total_amount')
            # If total includes GST (5%), calculate base
            # Total = Base * 1.05, so Base = Total / 1.05
            subtotal = round(total / 1.05, 2)
        
        return subtotal
    
    def _calculate_total_with_gst(self, data: Dict) -> float:
        """Calculate total amount with GST"""
        base = data.get('base_rent') or 0
        extra_km = data.get('extra_km_charge') or 0
        extra_hour = data.get('extra_hour_charge') or 0
        other = data.get('other_charges') or 0
        pickup_drop = data.get('pickup_drop_charges') or 0
        
        subtotal = base + extra_km + extra_hour + other + pickup_drop
        
        # Add 5% GST (2.5% CGST + 2.5% SGST)
        total_with_gst = subtotal * 1.05
        
        return round(total_with_gst, 2)
    
    def _format_booking_datetime(self, data: Dict) -> str:
        """Format booking date and time range with time included"""
        start = data.get('start_datetime')
        end = data.get('end_datetime')
        
        if start and end:
            # Convert from YYYY-MM-DD HH:MM to DD/MM/YYYY HH:MM format
            try:
                if 'T' in str(start):
                    # ISO format
                    start_dt = datetime.fromisoformat(str(start).replace('Z', '+00:00'))
                    end_dt = datetime.fromisoformat(str(end).replace('Z', '+00:00'))
                else:
                    # Standard format
                    start_dt = datetime.strptime(str(start), '%Y-%m-%d %H:%M')
                    end_dt = datetime.strptime(str(end), '%Y-%m-%d %H:%M')
                
                # Format with date and time
                return f"{start_dt.strftime('%d/%m/%Y %H:%M')} to {end_dt.strftime('%d/%m/%Y %H:%M')}"
            except:
                return f"{start} to {end}"
        elif start:
            try:
                if 'T' in str(start):
                    start_dt = datetime.fromisoformat(str(start).replace('Z', '+00:00'))
                else:
                    start_dt = datetime.strptime(str(start), '%Y-%m-%d %H:%M')
                return start_dt.strftime('%d/%m/%Y %H:%M')
            except:
                return str(start)
        
        return ""
    
    def _embed_document_images(self, ws, image_paths: List[str], start_row: int = 52):
        """
        Embed document images (Aadhaar, DL, etc.) below the invoice with MAXIMUM QUALITY
        
        Args:
            ws: Worksheet object
            image_paths: List of image file paths or bytes
            start_row: Starting row for images (default: 52, 4 rows gap after row 48)
        """
        if not image_paths:
            return
        
        current_row = start_row
        images_per_row = 2  # Show 2 images per row
        
        for idx, img_path in enumerate(image_paths):
            try:
                # Determine column (A or E for 2 images per row)
                col_offset = (idx % images_per_row) * 4  # 4 columns apart
                anchor_col = chr(65 + col_offset)  # A, E, I, etc.
                
                # Calculate row
                row_offset = idx // images_per_row
                anchor_row = current_row + (row_offset * 22)  # 22 rows per image
                
                # Load image
                if isinstance(img_path, bytes):
                    # Image is bytes
                    img_data = img_path
                    pil_img = Image.open(io.BytesIO(img_data))
                elif isinstance(img_path, str) and os.path.exists(img_path):
                    # Image is file path
                    pil_img = Image.open(img_path)
                else:
                    print(f"⚠️  Skipping invalid image: {img_path}")
                    continue
                
                # Convert to RGB if needed
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', pil_img.size, (255, 255, 255))
                    if pil_img.mode == 'P':
                        pil_img = pil_img.convert('RGBA')
                    if pil_img.mode == 'RGBA':
                        background.paste(pil_img, mask=pil_img.split()[-1])
                    else:
                        background.paste(pil_img)
                    pil_img = background
                elif pil_img.mode != 'RGB':
                    pil_img = pil_img.convert('RGB')
                
                # Get original dimensions
                original_width, original_height = pil_img.size
                
                # Target dimensions for Excel (larger = better quality)
                target_width = 800   # Increased from 600
                target_height = 600  # Increased from 450
                
                # Calculate aspect ratio
                aspect_ratio = original_width / original_height
                
                # Resize to target dimensions maintaining aspect ratio
                if aspect_ratio > 1:  # Wider than tall
                    new_width = target_width
                    new_height = int(target_width / aspect_ratio)
                else:  # Taller than wide
                    new_height = target_height
                    new_width = int(target_height * aspect_ratio)
                
                # Always resize to ensure consistent quality
                # Use HIGHEST quality resampling
                pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                
                # Save with MAXIMUM quality and explicit DPI
                img_byte_arr = io.BytesIO()
                
                # Save as PNG with HIGH DPI to prevent Excel scaling
                # Using 300 DPI (print quality) ensures Excel doesn't compress aggressively
                pil_img.save(
                    img_byte_arr, 
                    format='PNG',
                    compress_level=0,  # NO compression
                    optimize=False,
                    dpi=(300, 300)  # Print quality DPI - prevents Excel compression
                )
                img_byte_arr.seek(0)
                
                # Create Excel image
                xl_img = XLImage(img_byte_arr)
                
                # CRITICAL: Set explicit dimensions to prevent Excel auto-scaling
                # Convert pixels to Excel units (pixels * 0.75 = points)
                xl_img.width = new_width * 0.75
                xl_img.height = new_height * 0.75
                
                # Set anchor position
                anchor = f"{anchor_col}{anchor_row}"
                xl_img.anchor = anchor
                
                # Add to worksheet
                ws.add_image(xl_img)
                
                print(f"✅ Embedded image {idx + 1} at {anchor}")
                
            except Exception as e:
                print(f"⚠️  Failed to embed image {idx + 1}: {e}")
                continue
    
    def _generate_invoice_number(self) -> str:
        """Generate sequential invoice number in HD/YYYY-YY/XXX format"""
        import json
        import os
        
        counter_file = 'invoice_counter.json'
        
        # Load counter
        if os.path.exists(counter_file):
            with open(counter_file, 'r') as f:
                counter_data = json.load(f)
        else:
            # Initialize counter
            counter_data = {
                'last_invoice_number': 35,
                'financial_year': '2025-26'
            }
        
        # Get current financial year
        now = datetime.now()
        year = now.year
        next_year = (year + 1) % 100
        current_fy = f"{year}-{next_year:02d}"
        
        # Check if financial year changed
        if counter_data['financial_year'] != current_fy:
            # Reset counter for new financial year
            counter_data['financial_year'] = current_fy
            counter_data['last_invoice_number'] = 0
        
        # Increment counter
        counter_data['last_invoice_number'] += 1
        invoice_num = counter_data['last_invoice_number']
        
        # Save counter
        with open(counter_file, 'w') as f:
            json.dump(counter_data, f, indent=2)
        
        # Format: HD/2025-26/036
        return f"HD/{current_fy}/{invoice_num:03d}"


# Integration with existing extractor
def process_booking_to_excel(ocr_text: str, user_text: str, 
                             template_path: str = 'inn sample.xlsx',
                             output_path: str = 'filled_invoice.xlsx') -> Dict[str, Any]:
    """
    Complete pipeline: OCR/Text → JSON → Excel
    
    Args:
        ocr_text: Text extracted from image
        user_text: User-typed booking details
        template_path: Path to Excel template
        output_path: Path to save filled Excel
        
    Returns:
        Dictionary with extracted data and file path
    """
    from implementation_example import BookingDataExtractor
    
    # Extract data
    extractor = BookingDataExtractor()
    data = extractor.extract(ocr_text, user_text)
    
    # Write to Excel
    writer = HillDriveExcelWriter(template_path)
    output_file = writer.write(data, output_path)
    
    return {
        'extracted_data': data,
        'output_file': output_file,
        'confidence': data.get('extraction_confidence'),
        'calculation_verified': data.get('calculation_verified')
    }


# Example usage
if __name__ == '__main__':
    # Sample data
    sample_data = {
        'customer_name': 'Buen Manejo Del Campo India Pvt. Ltd.',
        'mobile_number': '8889302969',
        'address': 'Office no.4, 2nd Floor, Anmol Pride, Baner, Pune - 411045',
        'vehicle_name': 'Swift Dzire',
        'vehicle_number': 'RJ14AB1234',
        'start_datetime': '2026-01-25 07:00',
        'end_datetime': '2026-01-31 07:00',
        'duration_days': 6,
        'base_rent': 16200,
        'included_km': 600,
        'extra_km': 551,
        'extra_km_rate': 8,
        'extra_km_charge': 4408,
        'total_amount': 20608,
        'advance_paid': 10000,
        'place_of_supply': 'Udaipur',
        'delivery_address': 'Udaipur hub'
    }
    
    # Write to Excel
    writer = HillDriveExcelWriter('inn sample.xlsx')
    output = writer.write(sample_data, 'test_output.xlsx')
    print(f"✅ Invoice created: {output}")
